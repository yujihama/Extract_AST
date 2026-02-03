# %%
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
import base64
from typing import Any, Iterable, Optional
from src.utils import build_llm
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from pydantic import BaseModel, Field
import dotenv
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import zipfile
import shutil
import os
from xml.etree import ElementTree as ET

dotenv.load_dotenv()

# %%
def render_pdf_to_png_pages(
    pdf_path: Path,
    *,
    output_dir: Path,
    dpi: int = 150,
) -> list[Path]:
    """
    PDFをページごとにPNGへレンダリングして保存し、出力PNGパスのリストを返す。

    NOTE:
    - openpyxl はPDFを直接貼れないため、画像化して貼り付ける。
    - 依存: PyMuPDF（pymupdf）
    """
    import fitz  # PyMuPDF

    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    doc = fitz.open(str(pdf_path))
    out_paths: list[Path] = []
    try:
        for i in range(doc.page_count):
            page = doc.load_page(i)
            pix = page.get_pixmap(dpi=int(dpi))
            out_path = output_dir / f"{pdf_path.stem}__p{i+1:03}.png"
            pix.save(str(out_path))
            out_paths.append(out_path)
    finally:
        doc.close()

    return out_paths


# %%
# --- Excelシェイプオーバーレイ機能 ---
# 名前空間定義
_EXCEL_DRAWING_NAMESPACES = {
    '': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}

# 名前空間を登録
for _prefix, _uri in _EXCEL_DRAWING_NAMESPACES.items():
    ET.register_namespace(_prefix, _uri)


def _add_rectangle_shape_to_drawing_xml(
    xml_content: str,
    shape_id: int,
    name: str,
    from_col: int,
    from_col_off: int,
    from_row: int,
    from_row_off: int,
    to_col: int,
    to_col_off: int,
    to_row: int,
    to_row_off: int,
    line_width_emu: int = 38100,
    line_color: str = 'FF0000',
) -> str:
    """
    drawing1.xmlにrectangleシェイプを追加する。

    Args:
        xml_content: 元のdrawing XMLコンテンツ
        shape_id: シェイプID（ユニーク）
        name: シェイプ名
        from_col, from_col_off, from_row, from_row_off: 開始位置（列、列オフセットEMU、行、行オフセットEMU）
        to_col, to_col_off, to_row, to_row_off: 終了位置
        line_width_emu: 線の太さ（EMU単位、9525 = 約0.75pt, 38100 = 約3pt）
        line_color: 線の色（RGB hex、例: 'FF0000'=赤）

    Returns:
        更新されたXMLコンテンツ
    """
    root = ET.fromstring(xml_content)
    wsdr_ns = _EXCEL_DRAWING_NAMESPACES['']
    a_ns = _EXCEL_DRAWING_NAMESPACES['a']

    # twoCellAnchorを作成
    twoCellAnchor = ET.SubElement(root, '{%s}twoCellAnchor' % wsdr_ns)
    twoCellAnchor.set('editAs', 'oneCell')

    # from
    from_elem = ET.SubElement(twoCellAnchor, '{%s}from' % wsdr_ns)
    ET.SubElement(from_elem, '{%s}col' % wsdr_ns).text = str(from_col)
    ET.SubElement(from_elem, '{%s}colOff' % wsdr_ns).text = str(from_col_off)
    ET.SubElement(from_elem, '{%s}row' % wsdr_ns).text = str(from_row)
    ET.SubElement(from_elem, '{%s}rowOff' % wsdr_ns).text = str(from_row_off)

    # to
    to_elem = ET.SubElement(twoCellAnchor, '{%s}to' % wsdr_ns)
    ET.SubElement(to_elem, '{%s}col' % wsdr_ns).text = str(to_col)
    ET.SubElement(to_elem, '{%s}colOff' % wsdr_ns).text = str(to_col_off)
    ET.SubElement(to_elem, '{%s}row' % wsdr_ns).text = str(to_row)
    ET.SubElement(to_elem, '{%s}rowOff' % wsdr_ns).text = str(to_row_off)

    # sp (shape)
    sp = ET.SubElement(twoCellAnchor, '{%s}sp' % wsdr_ns)

    # nvSpPr
    nvSpPr = ET.SubElement(sp, '{%s}nvSpPr' % wsdr_ns)
    cNvPr = ET.SubElement(nvSpPr, '{%s}cNvPr' % wsdr_ns)
    cNvPr.set('id', str(shape_id))
    cNvPr.set('name', name)
    ET.SubElement(nvSpPr, '{%s}cNvSpPr' % wsdr_ns)

    # spPr
    spPr = ET.SubElement(sp, '{%s}spPr' % wsdr_ns)

    # prstGeom (rect)
    prstGeom = ET.SubElement(spPr, '{%s}prstGeom' % a_ns)
    prstGeom.set('prst', 'rect')
    ET.SubElement(prstGeom, '{%s}avLst' % a_ns)

    # noFill
    ET.SubElement(spPr, '{%s}noFill' % a_ns)

    # ln (line)
    ln = ET.SubElement(spPr, '{%s}ln' % a_ns)
    ln.set('w', str(line_width_emu))
    solidFill = ET.SubElement(ln, '{%s}solidFill' % a_ns)
    srgbClr = ET.SubElement(solidFill, '{%s}srgbClr' % a_ns)
    srgbClr.set('val', line_color)

    # clientData
    ET.SubElement(twoCellAnchor, '{%s}clientData' % wsdr_ns)

    return ET.tostring(root, encoding='unicode')


def _add_textbox_shape_to_drawing_xml(
    xml_content: str,
    shape_id: int,
    name: str,
    from_col: int,
    from_col_off: int,
    from_row: int,
    from_row_off: int,
    to_col: int,
    to_col_off: int,
    to_row: int,
    to_row_off: int,
    text: str,
    font_size_pt: int = 10,
    font_color: str = 'FF0000',
    fill_color: Optional[str] = 'FFFFCC',
    line_color: Optional[str] = '000000',
    line_width_emu: int = 9525,
) -> str:
    """
    drawing1.xmlにテキストボックスシェイプを追加する。

    Args:
        xml_content: 元のdrawing XMLコンテンツ
        shape_id: シェイプID（ユニーク）
        name: シェイプ名
        from_col, from_col_off, from_row, from_row_off: 開始位置
        to_col, to_col_off, to_row, to_row_off: 終了位置
        text: テキストボックス内のテキスト
        font_size_pt: フォントサイズ（ポイント）
        font_color: フォント色（RGB hex、例: 'FF0000'=赤）
        fill_color: 背景色（RGB hex、Noneで透明）
        line_color: 枠線色（RGB hex、Noneで枠線なし）
        line_width_emu: 枠線の太さ（EMU単位）

    Returns:
        更新されたXMLコンテンツ
    """
    root = ET.fromstring(xml_content)
    wsdr_ns = _EXCEL_DRAWING_NAMESPACES['']
    a_ns = _EXCEL_DRAWING_NAMESPACES['a']

    # twoCellAnchorを作成
    twoCellAnchor = ET.SubElement(root, '{%s}twoCellAnchor' % wsdr_ns)
    twoCellAnchor.set('editAs', 'oneCell')

    # from
    from_elem = ET.SubElement(twoCellAnchor, '{%s}from' % wsdr_ns)
    ET.SubElement(from_elem, '{%s}col' % wsdr_ns).text = str(from_col)
    ET.SubElement(from_elem, '{%s}colOff' % wsdr_ns).text = str(from_col_off)
    ET.SubElement(from_elem, '{%s}row' % wsdr_ns).text = str(from_row)
    ET.SubElement(from_elem, '{%s}rowOff' % wsdr_ns).text = str(from_row_off)

    # to
    to_elem = ET.SubElement(twoCellAnchor, '{%s}to' % wsdr_ns)
    ET.SubElement(to_elem, '{%s}col' % wsdr_ns).text = str(to_col)
    ET.SubElement(to_elem, '{%s}colOff' % wsdr_ns).text = str(to_col_off)
    ET.SubElement(to_elem, '{%s}row' % wsdr_ns).text = str(to_row)
    ET.SubElement(to_elem, '{%s}rowOff' % wsdr_ns).text = str(to_row_off)

    # sp (shape)
    sp = ET.SubElement(twoCellAnchor, '{%s}sp' % wsdr_ns)
    sp.set('macro', '')
    sp.set('textlink', '')

    # nvSpPr
    nvSpPr = ET.SubElement(sp, '{%s}nvSpPr' % wsdr_ns)
    cNvPr = ET.SubElement(nvSpPr, '{%s}cNvPr' % wsdr_ns)
    cNvPr.set('id', str(shape_id))
    cNvPr.set('name', name)
    cNvSpPr = ET.SubElement(nvSpPr, '{%s}cNvSpPr' % wsdr_ns)
    cNvSpPr.set('txBox', '1')

    # spPr
    spPr = ET.SubElement(sp, '{%s}spPr' % wsdr_ns)

    # prstGeom (rect)
    prstGeom = ET.SubElement(spPr, '{%s}prstGeom' % a_ns)
    prstGeom.set('prst', 'rect')
    ET.SubElement(prstGeom, '{%s}avLst' % a_ns)

    # 背景色
    if fill_color:
        solidFill = ET.SubElement(spPr, '{%s}solidFill' % a_ns)
        srgbClr = ET.SubElement(solidFill, '{%s}srgbClr' % a_ns)
        srgbClr.set('val', fill_color)
    else:
        ET.SubElement(spPr, '{%s}noFill' % a_ns)

    # 枠線
    ln = ET.SubElement(spPr, '{%s}ln' % a_ns)
    ln.set('w', str(line_width_emu))
    if line_color:
        solidFillLn = ET.SubElement(ln, '{%s}solidFill' % a_ns)
        srgbClrLn = ET.SubElement(solidFillLn, '{%s}srgbClr' % a_ns)
        srgbClrLn.set('val', line_color)
    else:
        ET.SubElement(ln, '{%s}noFill' % a_ns)

    # txBody (テキスト本体)
    txBody = ET.SubElement(sp, '{%s}txBody' % wsdr_ns)

    # bodyPr
    bodyPr = ET.SubElement(txBody, '{%s}bodyPr' % a_ns)
    bodyPr.set('vertOverflow', 'clip')
    bodyPr.set('horzOverflow', 'clip')
    bodyPr.set('wrap', 'square')
    bodyPr.set('lIns', '91440')  # 左余白 (EMU)
    bodyPr.set('tIns', '45720')  # 上余白 (EMU)
    bodyPr.set('rIns', '91440')  # 右余白 (EMU)
    bodyPr.set('bIns', '45720')  # 下余白 (EMU)
    bodyPr.set('anchor', 't')  # 上揃え

    # lstStyle
    ET.SubElement(txBody, '{%s}lstStyle' % a_ns)

    # テキストを改行で分割して各行をパラグラフとして追加
    lines = text.split('\n')
    for line in lines:
        p = ET.SubElement(txBody, '{%s}p' % a_ns)
        pPr = ET.SubElement(p, '{%s}pPr' % a_ns)
        pPr.set('algn', 'l')  # 左揃え

        r = ET.SubElement(p, '{%s}r' % a_ns)
        rPr = ET.SubElement(r, '{%s}rPr' % a_ns)
        rPr.set('lang', 'ja-JP')
        rPr.set('sz', str(font_size_pt * 100))  # フォントサイズ（100分の1ポイント）
        rPr.set('b', '1')  # 太字

        # フォント色
        solidFillText = ET.SubElement(rPr, '{%s}solidFill' % a_ns)
        srgbClrText = ET.SubElement(solidFillText, '{%s}srgbClr' % a_ns)
        srgbClrText.set('val', font_color)

        # フォント指定
        latin = ET.SubElement(rPr, '{%s}latin' % a_ns)
        latin.set('typeface', 'Meiryo UI')
        ea = ET.SubElement(rPr, '{%s}ea' % a_ns)
        ea.set('typeface', 'Meiryo UI')

        t = ET.SubElement(r, '{%s}t' % a_ns)
        t.text = line

    # clientData
    ET.SubElement(twoCellAnchor, '{%s}clientData' % wsdr_ns)

    return ET.tostring(root, encoding='unicode')


def _add_shapes_to_excel(
    excel_path: Path,
    shapes: list[dict],
    output_path: Optional[Path] = None,
    textboxes: Optional[list[dict]] = None,
) -> Path:
    """
    Excelファイルにシェイプを追加する（XMLを直接操作）。

    Args:
        excel_path: 入力Excelファイルパス
        shapes: シェイプ情報のリスト。各シェイプは以下のキーを持つdict:
            - from_col, from_col_off, from_row, from_row_off: 開始位置
            - to_col, to_col_off, to_row, to_row_off: 終了位置
            - line_width_emu: 線の太さ（EMU、オプション、デフォルト38100）
            - line_color: 線の色（RGB hex、オプション、デフォルト'FF0000'）
        output_path: 出力パス（Noneの場合は入力ファイルを上書き）
        textboxes: テキストボックス情報のリスト。各テキストボックスは以下のキーを持つdict:
            - from_col, from_col_off, from_row, from_row_off: 開始位置
            - to_col, to_col_off, to_row, to_row_off: 終了位置
            - text: テキスト内容
            - font_size_pt: フォントサイズ（オプション、デフォルト10）
            - font_color: フォント色（オプション、デフォルト'FF0000'）
            - fill_color: 背景色（オプション、デフォルト'FFFFCC'）
            - line_color: 枠線色（オプション、デフォルト'000000'）

    Returns:
        出力Excelファイルパス
    """
    excel_path = Path(excel_path)
    output_path = Path(output_path) if output_path else excel_path
    textboxes = textboxes or []

    if not shapes and not textboxes:
        if output_path != excel_path:
            shutil.copy(excel_path, output_path)
        return output_path

    # 一時ディレクトリに展開
    temp_dir = excel_path.parent / f".{excel_path.stem}_temp_shape"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)

    with zipfile.ZipFile(excel_path, 'r') as zf:
        zf.extractall(temp_dir)

    # drawing1.xmlを編集
    drawing_path = temp_dir / 'xl' / 'drawings' / 'drawing1.xml'
    if drawing_path.exists():
        with open(drawing_path, 'r', encoding='utf-8') as f:
            xml_content = f.read()

        # 各シェイプを追加
        for i, shape in enumerate(shapes):
            xml_content = _add_rectangle_shape_to_drawing_xml(
                xml_content,
                shape_id=1000 + i,  # 他のオブジェクトとかぶらないID
                name=f'Annotation Rectangle {i+1}',
                from_col=shape['from_col'],
                from_col_off=shape.get('from_col_off', 0),
                from_row=shape['from_row'],
                from_row_off=shape.get('from_row_off', 0),
                to_col=shape['to_col'],
                to_col_off=shape.get('to_col_off', 0),
                to_row=shape['to_row'],
                to_row_off=shape.get('to_row_off', 0),
                line_width_emu=shape.get('line_width_emu', 38100),
                line_color=shape.get('line_color', 'FF0000'),
            )

        # 各テキストボックスを追加
        for j, textbox in enumerate(textboxes):
            xml_content = _add_textbox_shape_to_drawing_xml(
                xml_content,
                shape_id=2000 + j,  # シェイプとかぶらないID
                name=f'Annotation Textbox {j+1}',
                from_col=textbox['from_col'],
                from_col_off=textbox.get('from_col_off', 0),
                from_row=textbox['from_row'],
                from_row_off=textbox.get('from_row_off', 0),
                to_col=textbox['to_col'],
                to_col_off=textbox.get('to_col_off', 0),
                to_row=textbox['to_row'],
                to_row_off=textbox.get('to_row_off', 0),
                text=textbox['text'],
                font_size_pt=textbox.get('font_size_pt', 10),
                font_color=textbox.get('font_color', 'FF0000'),
                fill_color=textbox.get('fill_color', 'FFFFCC'),
                line_color=textbox.get('line_color', '000000'),
            )

        with open(drawing_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)

    # ZIPに再圧縮
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = Path(root) / file
                arc_name = file_path.relative_to(temp_dir)
                zf.write(file_path, arc_name)

    # 一時ディレクトリを削除
    shutil.rmtree(temp_dir)
    return output_path


def _pixel_y_to_excel_row(
    pixel_y: int,
    image_start_row: int,
    image_height_px: int,
    rows_for_image: int,
) -> int:
    """
    画像内のピクセルY座標からExcelの行番号を計算する。

    Args:
        pixel_y: 画像内のY座標（ピクセル、上端=0）
        image_start_row: 画像が配置されている開始行（1-indexed）
        image_height_px: 画像の高さ（ピクセル）
        rows_for_image: 画像が占めるExcel行数

    Returns:
        対応するExcel行番号（1-indexed）
    """
    if image_height_px <= 0:
        return image_start_row
    # 比率でマッピング
    ratio = pixel_y / image_height_px
    row_offset = int(ratio * rows_for_image)
    return image_start_row + row_offset


def add_capture_sheet_with_images(
    excel_path: str,
    images_dir: str = "data/input/sample_image",
    row_height_pt: float = 15.0,
    margin_rows: int = 2,
    output_excel_path: Optional[str] = None,
    annotate_phrases: Optional[dict[str, list[str]]] = None,
    annotate_llm: Any = None,
    annotate_output_dir: str = "data/output/annotated_capture",
    output_excel_sheet_name: str = "キャプチャ",
    annotate_max_iters: int = 3,
    pdf_dpi: int = 150,
    pdf_render_dir: Optional[str] = None,
    annotation_column: str = "A",
    image_column: str = "B",
    max_workers: int = 3,
    use_shape_overlay: bool = False,
    use_textbox_annotation: bool = False,
) -> Path:
    """
    excel_path のExcelに sheet_name を作成し、images_dir 配下の画像をファイル名順に上から貼り付ける。
    既に同名シートがある場合は削除して作り直す。

    Args:
        use_shape_overlay: Trueの場合、画像に赤枠を埋め込む代わりに、
                          元画像をExcelに貼り付けてExcelの図形で赤枠をオーバーレイする。
                          LLMでの評価は既存通り画像に赤枠を埋め込んで行い、
                          確定後にExcel出力時に元画像+図形オーバーレイの方式に切り替える。
        use_textbox_annotation: Trueの場合、アノテーションをセルに直接書き込む代わりに、
                               テキストボックス図形として配置する。

    NOTE:
    - 画像貼り付けには Pillow が必要です（未インストールの場合は `pip install pillow`）。
    """
    excel_path_p = Path(excel_path)
    images_dir_p = Path(images_dir)

    if not excel_path_p.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path_p}")
    if not images_dir_p.exists():
        raise FileNotFoundError(f"画像フォルダが見つかりません: {images_dir_p}")

    # 対応拡張子:
    # - 画像: png/jpg/jpeg/bmp/gif/webp/tif/tiff
    # - PDF: pdf（ページごとにPNGへ変換して貼り付け）
    image_exts = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tif", ".tiff"}
    pdf_paths = sorted(
        [p for p in images_dir_p.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"],
        key=lambda p: p.name,
    )

    # 画像とPDFページを「ファイル名順→PDFはページ順」で並べる
    items: list[tuple[tuple[str, int], Path]] = []
    for p in images_dir_p.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() in image_exts:
            items.append(((p.name, 0), p))

    if pdf_paths:
        render_base = (
            Path(pdf_render_dir)
            if pdf_render_dir
            else Path("data/output/pdf_pages")
        )
        for pdf in pdf_paths:
            rendered = render_pdf_to_png_pages(
                pdf,
                output_dir=render_base,
                dpi=pdf_dpi,
            )
            for page_idx, png_path in enumerate(rendered, start=1):
                items.append(((pdf.name, page_idx), png_path))

    image_paths = [p for _, p in sorted(items, key=lambda t: t[0])]
    if not image_paths:
        raise FileNotFoundError(f"画像/PDFが見つかりません: {images_dir_p}")

    wb = openpyxl.load_workbook(excel_path_p)
    if output_excel_sheet_name in wb.sheetnames:
        wb.remove(wb[output_excel_sheet_name])

    ws = wb.create_sheet(title=output_excel_sheet_name)
    ws.sheet_view.showGridLines = False
    # アノテーション列（A列）の幅を設定
    ws.column_dimensions[annotation_column].width = 40
    # 画像列（B列）の幅を設定
    ws.column_dimensions[image_column].width = 120

    # 画像を上から順に配置
    # かぶり防止のため、シート側のデフォルト行高を固定し、その前提で「画像高さ→必要行数」を見積もる。
    ws.sheet_format.defaultRowHeight = row_height_pt

    # --- 並列処理: LLMアノテーションを事前に並列実行 ---
    annotation_results: dict[Path, AnnotationResult] = {}
    if annotate_phrases and annotate_llm is not None:
        # アノテーション対象の画像を特定
        images_to_annotate = [
            img_p for img_p in image_paths
            if Path(img_p).exists() and annotate_phrases.get(img_p.name, [])
        ]

        def _process_single_image(img_p: Path) -> tuple[Path, AnnotationResult]:
            """1画像のアノテーション処理（並列実行用）"""
            result = annotate_image_with_llm_red_boxes(
                llm=annotate_llm,
                image_path=img_p,
                target_phrases=annotate_phrases.get(img_p.name, []),
                output_dir=Path(annotate_output_dir),
                max_iters=annotate_max_iters,
                embed_text_in_image=False,
            )
            return (img_p, result)

        # 並列実行（max_workers で並列数を制御）
        if images_to_annotate:
            effective_workers = min(max_workers, len(images_to_annotate))
            print(f"[並列処理] {len(images_to_annotate)}画像を{effective_workers}並列で処理中...")
            with ThreadPoolExecutor(max_workers=effective_workers) as executor:
                futures = {
                    executor.submit(_process_single_image, img_p): img_p
                    for img_p in images_to_annotate
                }
                for future in as_completed(futures):
                    img_p = futures[future]
                    try:
                        path, result = future.result()
                        annotation_results[path] = result
                        print(f"[完了] {path.name}")
                    except Exception as e:
                        print(f"[エラー] {img_p.name}: {e}")
            print(f"[並列処理] 完了")

    # --- Excelへの書き込み（直列処理） ---
    # シェイプオーバーレイ用の情報を収集
    shape_overlay_info: list[dict] = []  # シェイプ情報のリスト
    textbox_info: list[dict] = []  # テキストボックス情報のリスト

    row = 1
    for image_path in image_paths:
        if not Path(image_path).exists():
            # フォルダ内のファイルが途中で削除された等のケースはスキップ
            continue
        start_row = row
        img_path_for_excel = image_path
        annotation_result: Optional[AnnotationResult] = annotation_results.get(image_path)

        if annotation_result is not None:
            if use_shape_overlay and annotation_result.original_image_path:
                # シェイプオーバーレイモード：元画像を使用
                img_path_for_excel = annotation_result.original_image_path
            else:
                # 通常モード：赤枠付き画像を使用
                img_path_for_excel = annotation_result.image_path

        img = XLImage(str(img_path_for_excel))

        # 元画像のサイズを取得（アノテーション位置計算用）
        original_width = annotation_result.image_width if annotation_result else 0
        original_height = annotation_result.image_height if annotation_result else 0
        if original_height == 0:
            with Image.open(image_path) as pil_im:
                original_width, original_height = pil_im.size

        # 横幅が大きすぎる場合だけ縮小（貼り付け後の見栄え用）
        max_width_px = 900
        scale = 1.0
        if getattr(img, "width", None) and img.width > max_width_px:
            scale = max_width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

        # 画像をB列（image_column）に配置
        ws.add_image(img, f"{image_column}{row}")

        # 次の画像開始行を計算（px→pt を 96dpi 前提で換算）
        height_px = int(getattr(img, "height", 0) or 0)
        height_pt = (height_px * 72.0) / 96.0 if height_px > 0 else 300.0
        rows_needed = max(1, int((height_pt / row_height_pt) + margin_rows))

        # 行の高さを明示し、テンプレ側の行高差異によるズレを抑える
        for r in range(start_row, start_row + rows_needed):
            ws.row_dimensions[r].height = row_height_pt

        # アノテーションをA列（annotation_column）に配置
        if annotation_result and annotation_result.annotations:
            for ann in annotation_result.annotations:
                # ピクセルY座標からExcel行を計算（元画像の座標系を使用）
                pixel_y = ann.box.y1
                target_row = _pixel_y_to_excel_row(
                    pixel_y=pixel_y,
                    image_start_row=start_row,
                    image_height_px=original_height,
                    rows_for_image=rows_needed - margin_rows,  # マージン分を除く
                )
                # アノテーションテキストを作成
                annotation_text = ann.phrase
                if ann.reason:
                    annotation_text += f"\n({ann.reason})"

                if use_textbox_annotation:
                    # テキストボックスモード：テキストボックス情報を収集
                    EMU_PER_PIXEL = 9525
                    # アノテーション列のインデックス（A=0）
                    ann_col_idx = ord(annotation_column.upper()) - ord('A')
                    # テキストボックスの高さ（行数）を計算
                    # テキスト行数に応じて高さを調整
                    text_lines = annotation_text.count('\n') + 1
                    textbox_rows = max(3, text_lines + 1)  # 最低3行分の高さ
                    # テキストボックスの幅（EMU）：列幅を基準に設定
                    textbox_width_emu = int(250 * 9525)  # 約250px幅

                    textbox_info.append({
                        'from_col': ann_col_idx,
                        'from_col_off': 0,
                        'from_row': target_row - 1,  # 0-indexed
                        'from_row_off': 0,
                        'to_col': ann_col_idx,
                        'to_col_off': textbox_width_emu,
                        'to_row': target_row - 1 + textbox_rows,  # 0-indexed
                        'to_row_off': 0,
                        'text': annotation_text,
                        'font_size_pt': 10,
                        'font_color': 'FF0000',
                        'fill_color': 'FFFFCC',
                        'line_color': '000000',
                    })
                else:
                    # 通常モード：セルにテキストを設定
                    cell = ws[f"{annotation_column}{target_row}"]
                    cell.value = annotation_text
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )
                    # 赤色のフォントを設定
                    cell.font = Font(
                        color="FF0000",
                        bold=True,
                    )

                # シェイプオーバーレイモードの場合、シェイプ情報を収集
                if use_shape_overlay:
                    # ピクセル座標をExcelセル座標に変換
                    # 画像列（B列=列インデックス1）からの相対位置を計算
                    # EMU (English Metric Units): 1ピクセル ≒ 9525 EMU (96dpi基準)
                    EMU_PER_PIXEL = 9525

                    # 画像のスケーリングを考慮
                    box_x1 = int(ann.box.x1 * scale)
                    box_y1 = int(ann.box.y1 * scale)
                    box_x2 = int(ann.box.x2 * scale)
                    box_y2 = int(ann.box.y2 * scale)

                    # 画像列のインデックス（B=1, C=2, ...）
                    image_col_idx = ord(image_column.upper()) - ord('A')

                    # Y座標をExcel行に変換
                    from_row_idx = _pixel_y_to_excel_row(
                        pixel_y=box_y1,
                        image_start_row=start_row,
                        image_height_px=height_px,
                        rows_for_image=rows_needed - margin_rows,
                    ) - 1  # 0-indexed
                    to_row_idx = _pixel_y_to_excel_row(
                        pixel_y=box_y2,
                        image_start_row=start_row,
                        image_height_px=height_px,
                        rows_for_image=rows_needed - margin_rows,
                    ) - 1  # 0-indexed

                    # X座標をオフセットとして計算（同一セル内のオフセット）
                    from_col_off = box_x1 * EMU_PER_PIXEL
                    to_col_off = box_x2 * EMU_PER_PIXEL

                    # 行内のY位置オフセットを計算
                    row_height_emu = int(row_height_pt * 12700)  # 1pt = 12700 EMU
                    from_row_off = 0  # 簡易化：行の先頭から
                    to_row_off = 0

                    shape_overlay_info.append({
                        'from_col': image_col_idx,
                        'from_col_off': from_col_off,
                        'from_row': from_row_idx,
                        'from_row_off': from_row_off,
                        'to_col': image_col_idx,
                        'to_col_off': to_col_off,
                        'to_row': to_row_idx,
                        'to_row_off': to_row_off,
                        'line_color': 'FF0000',  # 赤
                        'line_width_emu': 38100,  # 約3pt
                    })

        row += rows_needed

    save_path = Path(output_excel_path) if output_excel_path else excel_path_p
    try:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
    except PermissionError as e:
        # in-place保存がロック等で失敗する場合に備え、別名保存にフォールバック
        fallback = save_path
        if save_path == excel_path_p:
            fallback = excel_path_p.with_name(f"{excel_path_p.stem}__capture.xlsx")
        try:
            fallback.parent.mkdir(parents=True, exist_ok=True)
            wb.save(fallback)
            save_path = fallback
        except Exception:
            raise PermissionError(
                f"Excelファイルの保存に失敗しました。Excelで '{excel_path_p.name}' を開いている場合は閉じてから再実行してください: {excel_path_p}"
            ) from e

    # シェイプオーバーレイモードまたはテキストボックスモードの場合、保存後に図形を追加
    if shape_overlay_info or textbox_info:
        if shape_overlay_info:
            print(f"[シェイプオーバーレイ] {len(shape_overlay_info)}個のシェイプを追加中...")
        if textbox_info:
            print(f"[テキストボックス] {len(textbox_info)}個のテキストボックスを追加中...")
        _add_shapes_to_excel(save_path, shape_overlay_info, save_path, textboxes=textbox_info)
        print(f"[図形追加] 完了")

    return save_path


def _image_suffix_to_mime_type(p: Path) -> str:
    ext = p.suffix.lower().lstrip(".")
    if ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    if ext == "png":
        return "image/png"
    if ext == "webp":
        return "image/webp"
    if ext in {"tif", "tiff"}:
        return "image/tiff"
    if ext == "gif":
        return "image/gif"
    if ext == "bmp":
        return "image/bmp"
    # fallback
    return "image/png"


def _encode_image_as_data_url(image_path: Path) -> str:
    mime = _image_suffix_to_mime_type(image_path)
    b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{b64}"


def _coerce_int(v: Any, default: int = 0) -> int:
    try:
        if isinstance(v, bool):
            return default
        if isinstance(v, (int, float)):
            return int(round(float(v)))
        s = str(v).strip()
        if s == "":
            return default
        return int(round(float(s)))
    except Exception:
        return default


def _clamp_box(box: dict[str, Any], width: int, height: int) -> Optional[dict[str, int]]:
    """
    box: {"x1":..,"y1":..,"x2":..,"y2":..} を想定
    """
    x1 = _coerce_int(box.get("x1"), 0)
    y1 = _coerce_int(box.get("y1"), 0)
    x2 = _coerce_int(box.get("x2"), 0)
    y2 = _coerce_int(box.get("y2"), 0)

    x1 = max(0, min(width - 1, x1))
    x2 = max(0, min(width - 1, x2))
    y1 = max(0, min(height - 1, y1))
    y2 = max(0, min(height - 1, y2))

    # 正規化
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1

    # 最小サイズ確保
    if x2 - x1 < 2 or y2 - y1 < 2:
        return None
    return {"x1": x1, "y1": y1, "x2": x2, "y2": y2}


def _draw_red_boxes_on_image(
    image_path: Path,
    boxes: Iterable[dict[str, int]],
    *,
    stroke_width: int = 6,
    padding_px: int = 6,
    annotations: Optional[list[str]] = None,
) -> Image.Image:
    """
    画像に赤いバウンディングボックスと注釈を描画する。
    
    Args:
        image_path: 画像パス
        boxes: バウンディングボックスのリスト [{"x1": int, "y1": int, "x2": int, "y2": int}, ...]
        stroke_width: 枠線の太さ
        padding_px: パディング（px）
        annotations: 各BBに対応する注釈テキストのリスト（Noneの場合は注釈なし）
    """
    im = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(im)
    
    # フォントの準備（日本語対応）
    font_size = 24  # 1.5倍のサイズ（16 * 1.5 = 24）
    font = None
    # Windows環境での日本語フォントを優先的に試す
    japanese_fonts = [
        ("C:/Windows/Fonts/msgothic.ttc", 0),  # MSゴシック（TTC、インデックス0）
        ("C:/Windows/Fonts/meiryo.ttc", 0),    # メイリオ（TTC、インデックス0）
        ("C:/Windows/Fonts/msmincho.ttc", 0),  # MS明朝（TTC、インデックス0）
        ("C:/Windows/Fonts/msgothic.ttf", None),  # MSゴシック（TTF版）
        ("C:/Windows/Fonts/meiryo.ttf", None),    # メイリオ（TTF版）
    ]
    for font_path, font_index in japanese_fonts:
        try:
            if font_index is not None:
                # TTCファイルの場合はインデックスを指定
                font = ImageFont.truetype(font_path, font_size, index=font_index)
            else:
                font = ImageFont.truetype(font_path, font_size)
            break
        except Exception:
            continue
    
    # 日本語フォントが見つからない場合は英語フォントを試す
    if font is None:
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("DejaVuSans.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()
    
    boxes_list = list(boxes)
    annotations_list = annotations if annotations else [None] * len(boxes_list)
    
    for idx, b in enumerate(boxes_list):
        x1 = max(0, b["x1"] - padding_px)
        y1 = max(0, b["y1"] - padding_px)
        x2 = min(im.width - 1, b["x2"] + padding_px)
        y2 = min(im.height - 1, b["y2"] + padding_px)
        # Pillowのrectangleはoutlineのwidth指定が使える
        draw.rectangle([x1, y1, x2, y2], outline=(255, 0, 0), width=stroke_width)
        
        # 注釈テキストを描画
        if idx < len(annotations_list) and annotations_list[idx]:
            annotation_text = annotations_list[idx]
            # 注釈の位置：BBの左上の少し上（上にスペースがない場合はBBの内側）
            text_x = x1
            text_y = max(0, y1 - 35)  # BBの上35px（フォントサイズが大きくなったので調整）
            
            # 改行で分割
            lines = annotation_text.split('\n')
            line_height = 27  # 行の高さ（px、フォントサイズ24に合わせて1.5倍）
            
            # 各行の幅と高さを計算
            max_text_width = 0
            total_text_height = len(lines) * line_height
            
            for line in lines:
                try:
                    bbox = draw.textbbox((text_x, text_y), line, font=font)
                    line_width = bbox[2] - bbox[0]
                except AttributeError:
                    # textbboxが使えない場合はtextsizeを使用（古いPillow）
                    line_width, _ = draw.textsize(line, font=font)
                max_text_width = max(max_text_width, line_width)
            
            # 背景を描画（半透明の白背景）
            bg_padding = 4
            bg_x1 = text_x - bg_padding
            bg_y1 = text_y - bg_padding
            bg_x2 = text_x + max_text_width + bg_padding
            bg_y2 = text_y + total_text_height + bg_padding
            
            # 背景が画像外に出ないように調整
            bg_x1 = max(0, bg_x1)
            bg_y1 = max(0, bg_y1)
            bg_x2 = min(im.width - 1, bg_x2)
            bg_y2 = min(im.height - 1, bg_y2)
            
            # 半透明の背景を描画（RGBAモードで描画してから合成）
            overlay = Image.new("RGBA", im.size, (0, 0, 0, 0))
            overlay_draw = ImageDraw.Draw(overlay)
            overlay_draw.rectangle(
                [bg_x1, bg_y1, bg_x2, bg_y2],
                fill=(255, 255, 255, 220),  # 半透明の白
                outline=(255, 0, 0, 255),  # 赤い枠線
                width=2,
            )
            im = Image.alpha_composite(im.convert("RGBA"), overlay).convert("RGB")
            draw = ImageDraw.Draw(im)
            
            # テキストを各行ごとに描画（赤色）
            current_y = text_y
            for line in lines:
                if line.strip():  # 空行はスキップ
                    draw.text((text_x, current_y), line, fill=(255, 0, 0), font=font)
                current_y += line_height
    
    return im


def _save_image_as_png(im: Image.Image, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    im.save(output_path, format="PNG")
    return output_path


def _add_pixel_ruler_overlay(
    im: Image.Image,
    *,
    band_px: int = 36,
    tick_step_px: int = 50,
    major_step_px: int = 200,
    grid_step_px: int = 50,
    grid_alpha: int = 40,
    major_grid_alpha: int = 130,
    font_size: int = 16,
    boxes: Optional[list[dict[str, int]]] = None,
) -> Image.Image:
    """
    画像に「ピクセル目盛り（定規）」を重ねる。

    - 画像サイズは変えない（座標系を維持）
    - 上下左右に半透明の帯を敷き、tickとラベルを描く
    - 薄いグリッドも重ねる（座標の“物差し”）
    - boxesが指定された場合、それらを完全に含む200px単位のグリッド線を赤くする
    """
    import math
    
    base = im.convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    w, h = base.size
    band_px = max(16, int(band_px))
    tick_step_px = max(10, int(tick_step_px))
    major_step_px = max(tick_step_px, int(major_step_px))
    grid_step_px = max(10, int(grid_step_px))
    grid_alpha = max(0, min(255, int(grid_alpha)))
    major_grid_alpha = max(0, min(255, int(major_grid_alpha)))
    font_size = max(10, int(font_size))

    # フォント（Windows想定で Arial を優先し、なければ既定フォント）
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except Exception:
        try:
            font = ImageFont.truetype("DejaVuSans.ttf", font_size)
        except Exception:
            font = ImageFont.load_default()
    
    # 大きなフォント（ボックス関連のメモリ用）
    try:
        large_font = ImageFont.truetype("arial.ttf", int(font_size * 1.5))
    except Exception:
        try:
            large_font = ImageFont.truetype("DejaVuSans.ttf", int(font_size * 1.5))
        except Exception:
            large_font = font

    # バウンディングボックスを完全に含む200px単位のグリッド線を特定
    highlighted_x_lines: set[int] = set()
    highlighted_y_lines: set[int] = set()
    if boxes:
        for box in boxes:
            x1 = box.get("x1", 0)
            y1 = box.get("y1", 0)
            x2 = box.get("x2", 0)
            y2 = box.get("y2", 0)
            # ボックスを完全に含む最小の200px単位のグリッド範囲を計算
            x_min = math.floor(x1 / major_step_px) * major_step_px
            x_max = math.ceil(x2 / major_step_px) * major_step_px
            y_min = math.floor(y1 / major_step_px) * major_step_px
            y_max = math.ceil(y2 / major_step_px) * major_step_px
            # 範囲内の200px単位のグリッド線を追加
            for x in range(x_min, x_max + 1, major_step_px):
                if 0 <= x < w:
                    highlighted_x_lines.add(x)
            for y in range(y_min, y_max + 1, major_step_px):
                if 0 <= y < h:
                    highlighted_y_lines.add(y)

    # 薄いグリッド（内容を隠さないよう薄く）
    for x in range(0, w, grid_step_px):
        a = major_grid_alpha if (x % major_step_px) == 0 else grid_alpha
        draw.line([(x, 0), (x, h - 1)], fill=(0, 0, 0, a), width=1)
    for y in range(0, h, grid_step_px):
        a = major_grid_alpha if (y % major_step_px) == 0 else grid_alpha
        draw.line([(0, y), (w - 1, y)], fill=(0, 0, 0, a), width=1)
    
    # バウンディングボックスを完全に含む200px単位のグリッド線を赤く描画
    for x in highlighted_x_lines:
        draw.line([(x, 0), (x, h - 1)], fill=(255, 0, 0, 200), width=2)
    for y in highlighted_y_lines:
        draw.line([(0, y), (w - 1, y)], fill=(255, 0, 0, 200), width=2)

    # 半透明の帯（上・下・左・右）
    band_fill = (255, 255, 255, 170)
    draw.rectangle([0, 0, w - 1, band_px], fill=band_fill)  # top
    draw.rectangle([0, h - band_px, w - 1, h - 1], fill=band_fill)  # bottom
    draw.rectangle([0, 0, band_px, h - 1], fill=band_fill)  # left
    draw.rectangle([w - band_px, 0, w - 1, h - 1], fill=band_fill)  # right

    # 目盛り（上/下：x、左/右：y）
    for x in range(0, w, tick_step_px):
        is_major = (x % major_step_px) == 0
        is_highlighted = x in highlighted_x_lines
        tick_len = band_px - 6 if is_major else (band_px // 2)
        tick_color = (255, 0, 0, 255) if is_highlighted else (0, 0, 0, 255)
        tick_width = 3 if is_highlighted else (2 if is_major else 1)
        # top
        y0 = band_px - tick_len
        draw.line([(x, y0), (x, band_px)], fill=tick_color, width=tick_width)
        # bottom
        y1 = h - band_px
        draw.line([(x, y1), (x, y1 + tick_len)], fill=tick_color, width=tick_width)
        if is_major:
            text_font = large_font if is_highlighted else font
            text_color = (255, 0, 0, 255) if is_highlighted else (0, 0, 0, 255)
            draw.text((x + 2, 2), str(x), fill=text_color, font=text_font)
            draw.text((x + 2, h - band_px + 2), str(x), fill=text_color, font=text_font)

    for y in range(0, h, tick_step_px):
        is_major = (y % major_step_px) == 0
        is_highlighted = y in highlighted_y_lines
        tick_len = band_px - 6 if is_major else (band_px // 2)
        tick_color = (255, 0, 0, 255) if is_highlighted else (0, 0, 0, 255)
        tick_width = 3 if is_highlighted else (2 if is_major else 1)
        # left
        x0 = band_px - tick_len
        draw.line([(x0, y), (band_px, y)], fill=tick_color, width=tick_width)
        # right
        x1 = w - band_px
        draw.line([(x1, y), (x1 + tick_len, y)], fill=tick_color, width=tick_width)
        if is_major:
            text_font = large_font if is_highlighted else font
            text_color = (255, 0, 0, 255) if is_highlighted else (0, 0, 0, 255)
            draw.text((2, y + 2), str(y), fill=text_color, font=text_font)
            draw.text((w - band_px + 2, y + 2), str(y), fill=text_color, font=text_font)

    # 原点ラベル
    draw.text((2, 2), "0,0", fill=(200, 0, 0, 255), font=font)
    draw.text((w - band_px + 2, h - band_px + 2), f"{w-1},{h-1}", fill=(200, 0, 0, 255), font=font)

    return Image.alpha_composite(base, overlay).convert("RGB")


def _save_ruler_image(
    *,
    image_path: Path,
    output_path: Path,
    band_px: int = 36,
    tick_step_px: int = 50,
    major_step_px: int = 200,
    grid_step_px: int = 50,
    grid_alpha: int = 40,
    major_grid_alpha: int = 130,
    font_size: int = 16,
    boxes: Optional[list[dict[str, int]]] = None,
) -> Path:
    im = Image.open(image_path).convert("RGB")
    ruled = _add_pixel_ruler_overlay(
        im,
        band_px=band_px,
        tick_step_px=tick_step_px,
        major_step_px=major_step_px,
        grid_step_px=grid_step_px,
        grid_alpha=grid_alpha,
        major_grid_alpha=major_grid_alpha,
        font_size=font_size,
        boxes=boxes,
    )
    return _save_image_as_png(ruled, output_path)


class BoundingBox(BaseModel):
    """画像のピクセル座標（左上原点）。"""

    x1: int = Field(..., description="left (px)")
    y1: int = Field(..., description="top (px)")
    x2: int = Field(..., description="right (px)")
    y2: int = Field(..., description="bottom (px)")


class AnnotationInfo(BaseModel):
    """アノテーション情報（Excelテキスト挿入用）。"""
    phrase: str = Field(..., description="検出した文言")
    reason: str = Field("", description="根拠")
    box: BoundingBox = Field(..., description="バウンディングボックス")


class AnnotationResult(BaseModel):
    """annotate_image_with_llm_red_boxes の戻り値。"""
    image_path: Path = Field(..., description="処理後の画像パス（赤枠付き）")
    original_image_path: Optional[Path] = Field(None, description="元画像パス（赤枠なし）")
    annotations: list[AnnotationInfo] = Field(default_factory=list, description="アノテーション情報のリスト")
    image_width: int = Field(0, description="画像の幅（px）")
    image_height: int = Field(0, description="画像の高さ（px）")

    class Config:
        arbitrary_types_allowed = True


class PhraseMatch(BaseModel):
    phrase: str = Field(..., description="検出した文言（target_phrasesのうち最も近いもの）")
    box: BoundingBox
    reason: str = Field("", description="根拠（任意）")


class FindPhraseBoundingBoxesOutput(BaseModel):
    contains: bool
    matches: list[PhraseMatch] = Field(default_factory=list)


class BoxAdjustment(BaseModel):
    """各バウンディングボックスの相対調整（50px単位）"""
    box_index: int  # どのBBを調整するか（0から始まるインデックス）
    dx1: int = 0  # x1の調整量（負の値=左に拡張、正の値=右に縮小）
    dy1: int = 0  # y1の調整量（負の値=上に拡張、正の値=下に縮小）
    dx2: int = 0  # x2の調整量（正の値=右に拡張、負の値=左に縮小）
    dy2: int = 0  # y2の調整量（正の値=下に拡張、負の値=上に縮小）


class VerifyBoxesOutput(BaseModel):
    ok: bool
    adjustments: list[BoxAdjustment] = Field(default_factory=list)


def _invoke_structured_output(llm: Any, schema: Any, messages: list[Any]) -> Any:
    """
    LangChainの structured output を使ってschemaに沿った出力を得る。
    実装差異（schemaを位置引数/kwで渡す等）を吸収する。
    """
    # 1) Pydanticモデルを直接渡す（OpenAI互換）
    try:
        structured_llm = llm.with_structured_output(schema, method="json_schema")
        return structured_llm.invoke(messages)
    except TypeError:
        pass

    # 2) schema= で渡す
    try:
        structured_llm = llm.with_structured_output(schema=schema, method="json_schema")
        return structured_llm.invoke(messages)
    except TypeError:
        pass

    # 3) JSON schema(dict)で渡す（戻り値はdict想定）
    try:
        json_schema = schema.model_json_schema() if hasattr(schema, "model_json_schema") else schema
        structured_llm = llm.with_structured_output(schema=json_schema, method="json_schema")
        return structured_llm.invoke(messages)
    except Exception as e:
        raise RuntimeError("structured output の初期化に失敗しました") from e


def llm_find_phrase_bounding_boxes(
    *,
    llm: Any,
    image_path: Path,
    target_phrases: list[str],
) -> dict[str, Any]:
    """
    画像内に target_phrases のいずれかが含まれるかをLLM(Vision)で判定し、
    含まれる場合は bbox（ピクセル座標）を返す。

    戻り値（期待）:
    {
      "contains": true/false,
      "matches": [{"phrase": "...", "box": {"x1":..,"y1":..,"x2":..,"y2":..}, "reason": "..."}]
    }
    """
    from langchain_core.messages import HumanMessage

    with Image.open(image_path) as im:
        width, height = im.size

    phrases_str = "\n".join([f"- {p}" for p in target_phrases])
    data_url = _encode_image_as_data_url(image_path)
    print(data_url)

    prompt = f"""
あなたは画像を見て特定の情報が含まれるバウンディングボックスを特定するタスクを行っています。
# 作業ステップ
Step1: 以下の情報が画像内に含まれているか確認してください。
- 対象情報:
{phrases_str}

Step2: Step1で特定した情報が含まれるバウンディングボックス（各座標50px単位で指定）を全て回答してください。

# 対象の画像についての情報
- 画像サイズ: width={width}, height={height}
- 画像には「ピクセル定規」と「薄いグリッド」が重ねられています（上下左右=目盛り、200pxごとに数値ラベル、グリッドは50px間隔）。
- 定規は座標系の把握のためのもので、対象文言そのものではありません。

# 出力形式
JSON schema:
{{
  "contains": boolean,
  "matches": [
    {{
      "phrase": string,
      "box": {{"x1": number, "y1": number, "x2": number, "y2": number}},
      "reason": string
    }}
  ]
}}

ルール:
- 座標は画像のピクセル座標（左上原点）で50px単位で返してください。
- 対象の情報がいずれも含まれていない場合は contains=false かつ matches=[]
- 対象の情報がいずれかが含まれている場合は contains=true かつ matchesに該当箇所をすべて入れる
- phraseは対象の情報のうち最も近いものを入れる
""".strip()

    print(prompt)

    msg = HumanMessage(
        content=[
            {"type": "text", "text": prompt},
            # OpenAI互換の形式（data URL）
            {"type": "image_url", "image_url": {"url": data_url}},
        ]
    )

    out = _invoke_structured_output(llm, FindPhraseBoundingBoxesOutput, [msg])
    if isinstance(out, BaseModel):
        print(out.model_dump())
        return out.model_dump()
    if isinstance(out, dict):
        print(out)
        return out
    raise ValueError(f"structured output の戻り値が想定外です: type={type(out).__name__}, content={out!r}")


def llm_verify_boxes_on_annotated_image(
    *,
    llm: Any,
    original_image_path: Path,
    current_boxes: list[dict[str, int]],
    boxed_image_path: Path,
    target_phrases: list[str],
) -> dict[str, Any]:
    """
    original と boxed を比較し、赤枠が適切かをLLM(Vision)で検証する。
    OKなら {"ok": true}、NGなら50px単位の相対調整を返す。
    """
    from langchain_core.messages import HumanMessage

    with Image.open(original_image_path) as im:
        width, height = im.size

    phrases_str = "\n".join([f"- {p}" for p in target_phrases])
    original_url = _encode_image_as_data_url(original_image_path)
    boxed_url = _encode_image_as_data_url(boxed_image_path)

    # 現在のボックス情報をインデックス付きで表示
    boxes_with_index = []
    for i, box in enumerate(current_boxes):
        boxes_with_index.append(f"  box_index={i}: x1={box['x1']}, y1={box['y1']}, x2={box['x2']}, y2={box['y2']}")
    boxes_info = "\n".join(boxes_with_index)

    prompt = f"""
あなたは画像を見て特定の情報が含まれるバウンディングボックスが正しく赤枠で囲われているかを評価するタスクを行っています。
# 作業ステップ
Step1: 以下の情報が画像内に含まれているか確認してください。
- 対象情報:
{phrases_str}

Step2: Step1で特定した情報が赤枠に全て囲われているかを評価してください。
※全ての情報が赤枠に入っていれば、余白が含まれていても厳密に指摘する必要はありません。

Step3: Step2で評価した結果、対象の一部分しか赤枠で囲えていない箇所は、50px単位で相対的な調整量を指定してください。

# 対象の画像についての情報
- 画像サイズ: width={width}, height={height}
- 画像には「ピクセル定規」と「薄いグリッド」が重ねられています（上下左右=目盛り、200pxごとに数値ラベル、グリッドは50px間隔）。
- 現在描画されている赤枠のバウンディングボックス（インデックス付き）:
{boxes_info}

# 出力形式
{{
  "ok": boolean,
  "adjustments": [
    {{
      "box_index": number,  // 調整対象のボックスのインデックス（0から始まる）
      "dx1": number,  // x1の調整量（負の値=左に拡張、正の値=右に縮小）
      "dy1": number,  // y1の調整量（負の値=上に拡張、正の値=下に縮小）
      "dx2": number,  // x2の調整量（正の値=右に拡張、負の値=左に縮小）
      "dy2": number   // y2の調整量（正の値=下に拡張、負の値=上に縮小）
    }}
  ]
}}

ルール:
- 十分に正しければ ok=true, adjustments=[]
- 不十分なら ok=false とし、調整が必要なボックスの相対調整量を adjustments に列挙
- 調整量は必ず50px単位で指定（例: -50, 0, 50, 100, -100 など）
- 修正不要なボックスは adjustments に含めない（調整が必要なボックスのみ列挙）
- 調整の例:
  - 左に50px拡張したい場合: dx1=-50
  - 右に100px拡張したい場合: dx2=100
  - 上に50px拡張したい場合: dy1=-50
  - 下を50px縮小したい場合: dy2=-50
""".strip()

    msg = HumanMessage(
        content=[
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": original_url}},
            {"type": "image_url", "image_url": {"url": boxed_url}},
        ]
    )

    out = _invoke_structured_output(llm, VerifyBoxesOutput, [msg])
    if isinstance(out, BaseModel):
        return out.model_dump()
    if isinstance(out, dict):
        return out
    raise ValueError(f"structured output の戻り値が想定外です: type={type(out).__name__}, content={out!r}")


def annotate_image_with_llm_red_boxes(
    *,
    llm: Any,
    image_path: Path,
    target_phrases: list[str],
    output_dir: Path,
    max_iters: int = 3,
    stroke_width: int = 6,
    padding_px: int = 6,
    embed_text_in_image: bool = False,
) -> AnnotationResult:
    """
    フロー:
    1) LLMで対象文言の有無を判定し、あればbboxを取得
    2) bboxに赤枠を描画してPNGとして保存
    3) LLMで赤枠の妥当性を再確認し、NGなら修正bboxで描き直し
    """
    image_path = Path(image_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with Image.open(image_path) as im:
        width, height = im.size

    # LLM入力用に「定規付き」画像を用意（座標系は変えない）
    ruler_original_path = output_dir / f"{image_path.stem}__ruler.png"
    _save_ruler_image(image_path=image_path, output_path=ruler_original_path)

    # 1) 検出（定規付き画像を入力にする）
    det = llm_find_phrase_bounding_boxes(llm=llm, image_path=ruler_original_path, target_phrases=target_phrases)
    if not det.get("contains"):
        return AnnotationResult(
            image_path=image_path,
            original_image_path=image_path,
            annotations=[],
            image_width=width,
            image_height=height,
        )

    boxes_raw = []
    annotations_raw = []  # 各BBに対応する注釈テキスト
    for m in det.get("matches", []):
        if isinstance(m, dict) and isinstance(m.get("box"), dict):
            b = _clamp_box(m["box"], width, height)
            if b is not None:
                boxes_raw.append(b)
                # 注釈テキストを作成（phraseとreasonを組み合わせ）
                phrase = m.get("phrase", "")
                reason = m.get("reason", "")
                if phrase and reason:
                    annotation = f"{phrase}\n({reason})"
                elif phrase:
                    annotation = phrase
                elif reason:
                    annotation = reason
                else:
                    annotation = ""
                annotations_raw.append(annotation)
    if not boxes_raw:
        return AnnotationResult(
            image_path=image_path,
            original_image_path=image_path,
            annotations=[],
            image_width=width,
            image_height=height,
        )

    # 描画→検証→修正ループ
    #
    # 目的:
    # - `ok=False` が続くときにどの段階で何が起きているか追えるよう、
    #   各イテレーションの「断面」画像をすべて保存する。
    #
    # 保存先:
    # - <stem>__boxed_iter01.png, <stem>__boxed_iter02.png, ...
    # - <stem>__boxed_final.png（最後の採用結果）
    current_boxes = boxes_raw
    current_annotations = annotations_raw  # 注釈情報も保持
    iters = max(1, int(max_iters))

    last_iter_path: Optional[Path] = None
    for i in range(iters):
        iter_path = output_dir / f"{image_path.stem}__boxed_iter{i+1:02}.png"
        # 検証ループ中は注釈なしで描画（ノイズを避けるため）
        boxed_im = _draw_red_boxes_on_image(
            image_path,
            current_boxes,
            stroke_width=stroke_width,
            padding_px=padding_px,
            annotations=None,  # 注釈なし
        )
        _save_image_as_png(boxed_im, iter_path)
        last_iter_path = iter_path

        # 検証用：定規を重ねた断面画像も保存（2回目以降はバウンディングボックス情報を渡す）
        iter_ruler_path = output_dir / f"{image_path.stem}__boxed_iter{i+1:02}__ruler.png"
        _save_ruler_image(image_path=iter_path, output_path=iter_ruler_path, boxes=current_boxes if i > 0 else None)

        ver = llm_verify_boxes_on_annotated_image(
            llm=llm,
            current_boxes=current_boxes,
            original_image_path=ruler_original_path,
            boxed_image_path=iter_ruler_path,
            target_phrases=target_phrases,
        )
        if bool(ver.get("ok")):
            # OKになった断面を最終として採用
            final_path = output_dir / f"{image_path.stem}__boxed_final.png"
            if final_path != iter_path:
                final_im = Image.open(iter_path).convert("RGB")
                _save_image_as_png(final_im, final_path)

            # 定規付きfinalも保存（バウンディングボックス情報を渡す）
            final_ruler_path = output_dir / f"{image_path.stem}__boxed_final__ruler.png"
            _save_ruler_image(image_path=final_path, output_path=final_ruler_path, boxes=current_boxes)

            # アノテーション情報を構築
            result_annotations = []
            for j, box in enumerate(current_boxes):
                annotation_text = current_annotations[j] if j < len(current_annotations) else ""
                # annotation_textは "phrase\n(reason)" の形式なのでパース
                parts = annotation_text.split('\n', 1)
                phrase = parts[0] if parts else ""
                reason = parts[1].strip("()") if len(parts) > 1 else ""
                result_annotations.append(AnnotationInfo(
                    phrase=phrase,
                    reason=reason,
                    box=BoundingBox(x1=box["x1"], y1=box["y1"], x2=box["x2"], y2=box["y2"]),
                ))

            # embed_text_in_image=Trueの場合は注釈付き画像を生成
            if embed_text_in_image:
                final_annotated_path = output_dir / f"{image_path.stem}__boxed_final_annotated.png"
                final_annotations_text = []
                for j, box in enumerate(current_boxes):
                    if j < len(current_annotations):
                        final_annotations_text.append(current_annotations[j])
                    else:
                        final_annotations_text.append("")
                final_annotated_im = _draw_red_boxes_on_image(
                    image_path,
                    current_boxes,
                    stroke_width=stroke_width,
                    padding_px=padding_px,
                    annotations=final_annotations_text,
                )
                _save_image_as_png(final_annotated_im, final_annotated_path)
                return AnnotationResult(
                    image_path=final_annotated_path,
                    original_image_path=image_path,
                    annotations=result_annotations,
                    image_width=width,
                    image_height=height,
                )

            return AnnotationResult(
                image_path=final_path,
                original_image_path=image_path,
                annotations=result_annotations,
                image_width=width,
                image_height=height,
            )

        # adjustmentsから相対調整を取得
        adjustments: list[dict[str, Any]] = []
        for adj in (ver.get("adjustments") or [])[:10]:
            if isinstance(adj, dict) and "box_index" in adj:
                adjustments.append(adj)
        if not adjustments:
            # 修正が取れないならこの断面を最終として採用
            final_path = output_dir / f"{image_path.stem}__boxed_final.png"
            if last_iter_path and final_path != last_iter_path:
                final_im = Image.open(last_iter_path).convert("RGB")
                _save_image_as_png(final_im, final_path)
            final_ruler_path = output_dir / f"{image_path.stem}__boxed_final__ruler.png"
            if final_path.exists():
                _save_ruler_image(image_path=final_path, output_path=final_ruler_path, boxes=current_boxes)

            # アノテーション情報を構築
            result_annotations = []
            for j, box in enumerate(current_boxes):
                annotation_text = current_annotations[j] if j < len(current_annotations) else ""
                parts = annotation_text.split('\n', 1)
                phrase = parts[0] if parts else ""
                reason = parts[1].strip("()") if len(parts) > 1 else ""
                result_annotations.append(AnnotationInfo(
                    phrase=phrase,
                    reason=reason,
                    box=BoundingBox(x1=box["x1"], y1=box["y1"], x2=box["x2"], y2=box["y2"]),
                ))

            # embed_text_in_image=Trueの場合は注釈付き画像を生成
            if embed_text_in_image:
                final_annotated_path = output_dir / f"{image_path.stem}__boxed_final_annotated.png"
                final_annotations_text = []
                for j, box in enumerate(current_boxes):
                    if j < len(current_annotations):
                        final_annotations_text.append(current_annotations[j])
                    else:
                        final_annotations_text.append("")
                final_annotated_im = _draw_red_boxes_on_image(
                    image_path,
                    current_boxes,
                    stroke_width=stroke_width,
                    padding_px=padding_px,
                    annotations=final_annotations_text,
                )
                _save_image_as_png(final_annotated_im, final_annotated_path)
                return AnnotationResult(
                    image_path=final_annotated_path,
                    original_image_path=image_path,
                    annotations=result_annotations,
                    image_width=width,
                    image_height=height,
                )

            return AnnotationResult(
                image_path=final_path,
                original_image_path=image_path,
                annotations=result_annotations,
                image_width=width,
                image_height=height,
            )

        # adjustmentsを適用: box_indexで指定されたボックスに相対調整を適用
        new_boxes = list(current_boxes)
        for adj in adjustments:
            box_idx = adj.get("box_index", -1)
            if 0 <= box_idx < len(new_boxes):
                # 相対調整を適用（50px単位）
                dx1 = adj.get("dx1", 0)
                dy1 = adj.get("dy1", 0)
                dx2 = adj.get("dx2", 0)
                dy2 = adj.get("dy2", 0)

                old_box = new_boxes[box_idx]
                adjusted_box = {
                    "x1": old_box["x1"] + dx1,
                    "y1": old_box["y1"] + dy1,
                    "x2": old_box["x2"] + dx2,
                    "y2": old_box["y2"] + dy2,
                }
                # clampして範囲内に収める
                clamped = _clamp_box(adjusted_box, width, height)
                if clamped is not None:
                    new_boxes[box_idx] = clamped

        current_boxes = new_boxes

    # ここまで来たら「max_iters回の検証でokにならなかった」。
    # 最後の修正案を反映した画像を final として保存（再検証はしない）。
    final_path = output_dir / f"{image_path.stem}__boxed_final.png"
    boxed_im = _draw_red_boxes_on_image(
        image_path,
        current_boxes,
        stroke_width=stroke_width,
        padding_px=padding_px,
        annotations=None,  # 赤枠のみ
    )
    _save_image_as_png(boxed_im, final_path)
    final_ruler_path = output_dir / f"{image_path.stem}__boxed_final__ruler.png"
    _save_ruler_image(image_path=final_path, output_path=final_ruler_path, boxes=current_boxes)

    # アノテーション情報を構築
    result_annotations = []
    for j, box in enumerate(current_boxes):
        annotation_text = current_annotations[j] if j < len(current_annotations) else ""
        parts = annotation_text.split('\n', 1)
        phrase = parts[0] if parts else ""
        reason = parts[1].strip("()") if len(parts) > 1 else ""
        result_annotations.append(AnnotationInfo(
            phrase=phrase,
            reason=reason,
            box=BoundingBox(x1=box["x1"], y1=box["y1"], x2=box["x2"], y2=box["y2"]),
        ))

    # embed_text_in_image=Trueの場合は注釈付き画像を生成
    if embed_text_in_image:
        final_annotated_path = output_dir / f"{image_path.stem}__boxed_final_annotated.png"
        final_annotations_text = []
        for j, box in enumerate(current_boxes):
            if j < len(current_annotations):
                final_annotations_text.append(current_annotations[j])
            else:
                final_annotations_text.append("")
        final_annotated_im = _draw_red_boxes_on_image(
            image_path,
            current_boxes,
            stroke_width=stroke_width,
            padding_px=padding_px,
            annotations=final_annotations_text,
        )
        _save_image_as_png(final_annotated_im, final_annotated_path)
        return AnnotationResult(
            image_path=final_annotated_path,
            original_image_path=image_path,
            annotations=result_annotations,
            image_width=width,
            image_height=height,
        )

    return AnnotationResult(
        image_path=final_path,
        original_image_path=image_path,
        annotations=result_annotations,
        image_width=width,
        image_height=height,
    )

# %%
def capture_insert_sheet(
    excel_path: str,
    images_dir: str = "data/input/sample_image",
    phrases_to_box: list[str] = [],
    row_height_pt: float = 15.0,
    margin_rows: int = 2,
    output_excel_path: Optional[str] = None,
    output_excel_sheet_name: str = "キャプチャ",
    max_workers: int = 3,
    use_shape_overlay: bool = False,
    use_textbox_annotation: bool = False,
) -> Path:
    """
    画像をExcelに貼り付け、LLMでアノテーションを検出してA列にテキストを挿入する。

    Args:
        excel_path: 入力Excelファイルパス
        images_dir: 画像フォルダ
        phrases_to_box: 画像ファイル名→検出対象フレーズのマッピング
        row_height_pt: 行の高さ（pt）
        margin_rows: 画像間のマージン行数
        output_excel_path: 出力Excelパス（Noneなら入力ファイルを上書き）
        output_excel_sheet_name: 出力シート名
        max_workers: 並列処理のワーカー数（デフォルト3）
        use_shape_overlay: Trueの場合、画像に赤枠を埋め込む代わりに、
                          元画像をExcelに貼り付けてExcelの図形で赤枠をオーバーレイする。
        use_textbox_annotation: Trueの場合、アノテーションをセルに直接書き込む代わりに、
                               テキストボックス図形として配置する。
    """
    llm_complex = build_llm(model="gpt-5.2")
    output_path = add_capture_sheet_with_images(
        excel_path,
        annotate_phrases=phrases_to_box,
        annotate_llm=llm_complex,  # gpt-5.2 を想定（Vision対応）
        annotate_output_dir="data/output/annotated_capture",
        annotate_max_iters=3,
        output_excel_path=output_excel_path,
        output_excel_sheet_name=output_excel_sheet_name,
        max_workers=max_workers,
        use_shape_overlay=use_shape_overlay,
        use_textbox_annotation=use_textbox_annotation,
    )
    print(f"キャプチャ付きExcelを出力しました: {output_path}")


if __name__ == "__main__":
    file_path = "data/input/複数シート型1_セクション別シート.xlsx"
    # 例: 画像内にこれらの文言が含まれていたら赤枠を付ける
    phrases_to_box = {
        "仕訳定義書__p001.png": [
            "<借⽅勘定科⽬についての特別ルール>",
        ],
        "invoice_pattern5_slight_texture.png": [
            "<合計金額を表す箇所>",
            "<押印が含まれる箇所>",
        ],
    }

    output_path = capture_insert_sheet(
        file_path,
        phrases_to_box=phrases_to_box,
        output_excel_path="data/output/複数シート型1_セクション別シート__capture.xlsx",
    )
    print(f"Completed!")

# %%
