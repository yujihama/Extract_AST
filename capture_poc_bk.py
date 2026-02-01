# %%
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage
import base64
import json
from typing import Any, Iterable, Optional
from src.utils import build_llm
import openpyxl
from PIL import Image, ImageDraw

# %%
def render_pdf_to_png_pages(
    pdf_path: Path,
    *,
    output_dir: Path,
    dpi: int = 150,
) -> list[Path]:
    """
    PDFをページごとにPNGへレンダリングして保存し、出力PNGパスのリストを返す。

    NOTE:
    - openpyxl はPDFを直接貼れないため、画像化して貼り付ける。
    - 依存: PyMuPDF（pymupdf）
    """
    import fitz  # PyMuPDF

    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    doc = fitz.open(str(pdf_path))
    out_paths: list[Path] = []
    try:
        for i in range(doc.page_count):
            page = doc.load_page(i)
            pix = page.get_pixmap(dpi=int(dpi))
            out_path = output_dir / f"{pdf_path.stem}__p{i+1:03}.png"
            pix.save(str(out_path))
            out_paths.append(out_path)
    finally:
        doc.close()

    return out_paths


def add_capture_sheet_with_images(
    excel_path: str,
    images_dir: str = "data/input/sample_image",
    sheet_name: str = "キャプチャ",
    row_height_pt: float = 15.0,
    margin_rows: int = 2,
    output_excel_path: Optional[str] = None,
    # 画像に赤枠を書き込む（方式A）
    annotate_phrases: Optional[list[str]] = None,
    annotate_llm: Any = None,
    annotate_output_dir: str = "data/output/annotated_capture",
    annotate_max_iters: int = 3,
    pdf_dpi: int = 150,
    pdf_render_dir: Optional[str] = None,
) -> Path:
    """
    excel_path のExcelに sheet_name を作成し、images_dir 配下の画像をファイル名順に上から貼り付ける。
    既に同名シートがある場合は削除して作り直す。

    NOTE:
    - 画像貼り付けには Pillow が必要です（未インストールの場合は `pip install pillow`）。
    """
    excel_path_p = Path(excel_path)
    images_dir_p = Path(images_dir)

    if not excel_path_p.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path_p}")
    if not images_dir_p.exists():
        raise FileNotFoundError(f"画像フォルダが見つかりません: {images_dir_p}")

    # 対応拡張子:
    # - 画像: png/jpg/jpeg/bmp/gif/webp/tif/tiff
    # - PDF: pdf（ページごとにPNGへ変換して貼り付け）
    image_exts = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tif", ".tiff"}
    pdf_paths = sorted(
        [p for p in images_dir_p.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"],
        key=lambda p: p.name,
    )

    # 画像とPDFページを「ファイル名順→PDFはページ順」で並べる
    items: list[tuple[tuple[str, int], Path]] = []
    for p in images_dir_p.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() in image_exts:
            items.append(((p.name, 0), p))

    if pdf_paths:
        render_base = (
            Path(pdf_render_dir)
            if pdf_render_dir
            else Path("data/output/pdf_pages")
        )
        for pdf in pdf_paths:
            rendered = render_pdf_to_png_pages(
                pdf,
                output_dir=render_base,
                dpi=pdf_dpi,
            )
            for page_idx, png_path in enumerate(rendered, start=1):
                items.append(((pdf.name, page_idx), png_path))

    image_paths = [p for _, p in sorted(items, key=lambda t: t[0])]
    if not image_paths:
        raise FileNotFoundError(f"画像/PDFが見つかりません: {images_dir_p}")

    wb = openpyxl.load_workbook(excel_path_p)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    # 画像を上から順に配置
    # かぶり防止のため、シート側のデフォルト行高を固定し、その前提で「画像高さ→必要行数」を見積もる。
    ws.sheet_format.defaultRowHeight = row_height_pt
    row = 1
    for image_path in image_paths:
        start_row = row
        img_path_for_excel = image_path
        if annotate_phrases and annotate_llm is not None:
            img_path_for_excel = annotate_image_with_llm_red_boxes(
                llm=annotate_llm,
                image_path=image_path,
                target_phrases=annotate_phrases,
                output_dir=Path(annotate_output_dir),
                max_iters=annotate_max_iters,
            )

        img = XLImage(str(img_path_for_excel))

        # 横幅が大きすぎる場合だけ縮小（貼り付け後の見栄え用）
        max_width_px = 900
        if getattr(img, "width", None) and img.width > max_width_px:
            scale = max_width_px / float(img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

        ws.add_image(img, f"A{row}")

        # 次の画像開始行を計算（px→pt を 96dpi 前提で換算）
        height_px = int(getattr(img, "height", 0) or 0)
        height_pt = (height_px * 72.0) / 96.0 if height_px > 0 else 300.0
        rows_needed = max(1, int((height_pt / row_height_pt) + margin_rows))

        # 行の高さを明示し、テンプレ側の行高差異によるズレを抑える
        for r in range(start_row, start_row + rows_needed):
            ws.row_dimensions[r].height = row_height_pt

        row += rows_needed

    save_path = Path(output_excel_path) if output_excel_path else excel_path_p
    try:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
    except PermissionError as e:
        # in-place保存がロック等で失敗する場合に備え、別名保存にフォールバック
        fallback = save_path
        if save_path == excel_path_p:
            fallback = excel_path_p.with_name(f"{excel_path_p.stem}__capture.xlsx")
        try:
            fallback.parent.mkdir(parents=True, exist_ok=True)
            wb.save(fallback)
            return fallback
        except Exception:
            raise PermissionError(
                f"Excelファイルの保存に失敗しました。Excelで '{excel_path_p.name}' を開いている場合は閉じてから再実行してください: {excel_path_p}"
            ) from e

    return save_path


def _image_suffix_to_mime_type(p: Path) -> str:
    ext = p.suffix.lower().lstrip(".")
    if ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    if ext == "png":
        return "image/png"
    if ext == "webp":
        return "image/webp"
    if ext in {"tif", "tiff"}:
        return "image/tiff"
    if ext == "gif":
        return "image/gif"
    if ext == "bmp":
        return "image/bmp"
    # fallback
    return "image/png"


def _encode_image_as_data_url(image_path: Path) -> str:
    mime = _image_suffix_to_mime_type(image_path)
    b64 = base64.b64encode(image_path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{b64}"


def _coerce_int(v: Any, default: int = 0) -> int:
    try:
        if isinstance(v, bool):
            return default
        if isinstance(v, (int, float)):
            return int(round(float(v)))
        s = str(v).strip()
        if s == "":
            return default
        return int(round(float(s)))
    except Exception:
        return default


def _clamp_box(box: dict[str, Any], width: int, height: int) -> Optional[dict[str, int]]:
    """
    box: {"x1":..,"y1":..,"x2":..,"y2":..} を想定
    """
    x1 = _coerce_int(box.get("x1"), 0)
    y1 = _coerce_int(box.get("y1"), 0)
    x2 = _coerce_int(box.get("x2"), 0)
    y2 = _coerce_int(box.get("y2"), 0)

    x1 = max(0, min(width - 1, x1))
    x2 = max(0, min(width - 1, x2))
    y1 = max(0, min(height - 1, y1))
    y2 = max(0, min(height - 1, y2))

    # 正規化
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1

    # 最小サイズ確保
    if x2 - x1 < 2 or y2 - y1 < 2:
        return None
    return {"x1": x1, "y1": y1, "x2": x2, "y2": y2}


def _draw_red_boxes_on_image(
    image_path: Path,
    boxes: Iterable[dict[str, int]],
    *,
    stroke_width: int = 6,
    padding_px: int = 6,
) -> Image.Image:
    im = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(im)
    for b in boxes:
        x1 = max(0, b["x1"] - padding_px)
        y1 = max(0, b["y1"] - padding_px)
        x2 = min(im.width - 1, b["x2"] + padding_px)
        y2 = min(im.height - 1, b["y2"] + padding_px)
        # Pillowのrectangleはoutlineのwidth指定が使える
        draw.rectangle([x1, y1, x2, y2], outline=(255, 0, 0), width=stroke_width)
    return im


def _save_image_as_png(im: Image.Image, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    im.save(output_path, format="PNG")
    return output_path


def _parse_json_from_text(text: str) -> Any:
    """
    LLMが前後に説明文を付けても拾えるように、最初のJSONオブジェクト/配列を雑に抽出する。
    """
    s = (text or "").strip()
    if not s:
        raise ValueError("LLM出力が空です")

    # まずはそのまま
    try:
        return json.loads(s)
    except Exception:
        pass

    # ```json ... ``` を剥がす
    if "```" in s:
        parts = s.split("```")
        # fence内を優先
        for part in parts:
            p = part.strip()
            if p.startswith("json"):
                p = p[4:].strip()
            if p.startswith("{") or p.startswith("["):
                try:
                    return json.loads(p)
                except Exception:
                    continue

    # 最初の { or [ から最後の } or ] まで
    for start_char, end_char in [("{", "}"), ("[", "]")]:
        i = s.find(start_char)
        j = s.rfind(end_char)
        if i != -1 and j != -1 and j > i:
            candidate = s[i : j + 1]
            try:
                return json.loads(candidate)
            except Exception:
                continue

    raise ValueError("LLM出力からJSONを抽出できませんでした")


def llm_find_phrase_bounding_boxes(
    *,
    llm: Any,
    image_path: Path,
    target_phrases: list[str],
) -> dict[str, Any]:
    """
    画像内に target_phrases のいずれかが含まれるかをLLM(Vision)で判定し、
    含まれる場合は bbox（ピクセル座標）を返す。

    戻り値（期待）:
    {
      "contains": true/false,
      "matches": [{"phrase": "...", "box": {"x1":..,"y1":..,"x2":..,"y2":..}, "reason": "..."}]
    }
    """
    from langchain_core.messages import HumanMessage

    with Image.open(image_path) as im:
        width, height = im.size

    phrases_str = "\n".join([f"- {p}" for p in target_phrases])
    data_url = _encode_image_as_data_url(image_path)

    prompt = f"""
あなたは画像を見て、特定の文言が含まれるかを判断する監査者です。
以下のいずれかの文言が画像内に含まれているか確認してください。

対象文言:
{phrases_str}

出力は必ずJSONのみ。座標は画像のピクセル座標（左上原点）で返してください。
画像サイズ: width={width}, height={height}

JSON schema:
{{
  "contains": boolean,
  "matches": [
    {{
      "phrase": string,
      "box": {{"x1": number, "y1": number, "x2": number, "y2": number}},
      "reason": string
    }}
  ]
}}

ルール:
- 該当なしなら contains=false かつ matches=[]
- 該当ありなら contains=true かつ matchesに該当箇所をすべて入れる（多くても最大10件）
- phraseは対象文言のうち最も近いものを入れる
""".strip()

    msg = HumanMessage(
        content=[
            {"type": "text", "text": prompt},
            # OpenAI互換の形式（data URL）
            {"type": "image_url", "image_url": {"url": data_url}},
        ]
    )
    res = llm.invoke([msg])
    data = _parse_json_from_text(getattr(res, "content", "") or "")

    if not isinstance(data, dict):
        raise ValueError("LLMのJSONがdictではありません")
    if "contains" not in data:
        raise ValueError("LLMのJSONに contains がありません")
    if "matches" not in data or not isinstance(data["matches"], list):
        data["matches"] = []
    print(data)
    return data


def llm_verify_boxes_on_annotated_image(
    *,
    llm: Any,
    original_image_path: Path,
    boxed_image_path: Path,
    target_phrases: list[str],
) -> dict[str, Any]:
    """
    original と boxed を比較し、赤枠が適切かをLLM(Vision)で検証する。
    OKなら {"ok": true}、NGなら修正版のboxesを返す。
    """
    from langchain_core.messages import HumanMessage

    phrases_str = "\n".join([f"- {p}" for p in target_phrases])
    original_url = _encode_image_as_data_url(original_image_path)
    boxed_url = _encode_image_as_data_url(boxed_image_path)

    prompt = f"""
あなたは画像注釈（赤枠）の検証者です。
1枚目が元画像、2枚目が赤枠付き画像です。
対象文言が含まれる場合、赤枠が対象文言を適切に囲っているか確認してください。

対象文言:
{phrases_str}

出力は必ずJSONのみ:
{{
  "ok": boolean,
  "notes": string,
  "corrected_boxes": [
    {{"x1": number, "y1": number, "x2": number, "y2": number}}
  ]
}}

ルール:
- 十分に正しければ ok=true, corrected_boxes=[]
- 不十分なら ok=false とし、正しい赤枠の座標（元画像のピクセル座標）を corrected_boxes に列挙（最大10件）
""".strip()

    msg = HumanMessage(
        content=[
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": original_url}},
            {"type": "image_url", "image_url": {"url": boxed_url}},
        ]
    )
    res = llm.invoke([msg])
    data = _parse_json_from_text(getattr(res, "content", "") or "")
    if not isinstance(data, dict) or "ok" not in data:
        raise ValueError("LLM検証のJSON形式が不正です")
    if "corrected_boxes" not in data or not isinstance(data["corrected_boxes"], list):
        data["corrected_boxes"] = []
    print(data)
    return data


def annotate_image_with_llm_red_boxes(
    *,
    llm: Any,
    image_path: Path,
    target_phrases: list[str],
    output_dir: Path,
    max_iters: int = 3,
    stroke_width: int = 6,
    padding_px: int = 6,
) -> Path:
    """
    フロー:
    1) LLMで対象文言の有無を判定し、あればbboxを取得
    2) bboxに赤枠を描画してPNGとして保存
    3) LLMで赤枠の妥当性を再確認し、NGなら修正bboxで描き直し
    """
    image_path = Path(image_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with Image.open(image_path) as im:
        width, height = im.size

    # 1) 検出
    det = llm_find_phrase_bounding_boxes(llm=llm, image_path=image_path, target_phrases=target_phrases)
    if not det.get("contains"):
        return image_path

    boxes_raw = []
    for m in det.get("matches", [])[:10]:
        if isinstance(m, dict) and isinstance(m.get("box"), dict):
            b = _clamp_box(m["box"], width, height)
            if b is not None:
                boxes_raw.append(b)
    if not boxes_raw:
        return image_path

    # 描画→検証→修正ループ
    current_boxes = boxes_raw
    boxed_path = output_dir / f"{image_path.stem}__boxed.png"
    for _ in range(max(1, int(max_iters))):
        boxed_im = _draw_red_boxes_on_image(
            image_path,
            current_boxes,
            stroke_width=stroke_width,
            padding_px=padding_px,
        )
        _save_image_as_png(boxed_im, boxed_path)

        ver = llm_verify_boxes_on_annotated_image(
            llm=llm,
            original_image_path=image_path,
            boxed_image_path=boxed_path,
            target_phrases=target_phrases,
        )
        if bool(ver.get("ok")):
            return boxed_path

        corrected = []
        for b in (ver.get("corrected_boxes") or [])[:10]:
            if isinstance(b, dict):
                cb = _clamp_box(b, width, height)
                if cb is not None:
                    corrected.append(cb)
        if not corrected:
            # 修正が取れないなら現状を採用
            return boxed_path
        current_boxes = corrected

    return boxed_path

# %%
# 実行（必要に応じてコメントアウトしてください）
# NOTE: 直接実行時のみ動かす（import時に重い処理が走らないようにする）
if __name__ == "__main__":
    file_path = "data/input/複数シート型1_セクション別シート.xlsx"
    # 例: 画像内にこれらの文言が含まれていたら赤枠を付ける
    phrases_to_box = [
        "CONFIDENTIAL",
        "機密",
        "社外秘",
        "承認",
        "決裁",
        f"<報酬金額にあたる具体的な金額>",
    ]
    llm_complex = build_llm(model="gpt-5.2")
    output_path = add_capture_sheet_with_images(
        file_path,
        annotate_phrases=phrases_to_box,
        annotate_llm=llm_complex,  # gpt-5.2 を想定（Vision対応）
        annotate_output_dir="data/output/annotated_capture",
        annotate_max_iters=3,
        output_excel_path="data/output/複数シート型1_セクション別シート__capture.xlsx",
    )
    print(f"キャプチャ付きExcelを出力しました: {output_path}")

# %%
