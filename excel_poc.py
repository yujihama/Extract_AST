# %%
from langchain_core.callbacks import LLMManagerMixin
import pandas as pd
import openpyxl
import os
import dotenv
from langchain.agents import create_agent
from langchain_core.tools import tool
from src.utils import build_llm, print_message_logs, extract_message_logs
from deepagents.middleware import SubAgentMiddleware

# %%
dotenv.load_dotenv()

# %%
llm = build_llm()
llm_complex = build_llm(model="gpt-5.2")

# %%
# tools
# 1. read_excel
@tool
def read_excel(file_path: str) -> str:
    """
    エクセルを読み込み、全シートを対象に以下の変換を行ったテキストを返却する。
    ・各シートごとに「=== Sheet: シート名 ===」というヘッダーを付与
    ・セルを「|」で区切る
    ・各行の先頭に「row_1:」「row_2:」というように行番号を付与する
    ・「XXXX」と記載されているセルは更新すべきセルであることが分かるように「Need_to_update」と置換する。
      その際セルアドレスを付与する。
      - XXXXが1シートのみにある場合: (A1)Need_to_update
      - XXXXが複数シートにある場合: (Sheet1!A1)Need_to_update
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        Text with converted cells from all sheets
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Step1: XXXXセルがあるシートを特定
        sheets_with_xxxx = set()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if any(cell == "XXXX" for cell in row if cell is not None):
                    sheets_with_xxxx.add(sheet_name)
                    break
        
        # 複数シートにXXXXがあるかどうかでフォーマット決定
        use_sheet_prefix = len(sheets_with_xxxx) > 1
        
        # Step2: 全シートのテキストを生成
        all_sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            rows.append(f"=== Sheet: {sheet_name} ===")
            
            for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_text = []
                for col_idx, cell_value in enumerate(row, start=1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # セルアドレスの形式を決定
                    if use_sheet_prefix:
                        cell_addr = f"{sheet_name}!{col_letter}{row_num}"
                    else:
                        cell_addr = f"{col_letter}{row_num}"
                    
                    value = str(cell_value) if cell_value is not None else ""
                    if value == "XXXX":
                        value = f"({cell_addr})Need_to_update"
                    row_text.append(value)
                
                row_line = f"row_{row_num}:" + "|".join(row_text)
                rows.append(row_line)
            
            all_sheets_text.append("\n".join(rows))
        
        return "\n\n".join(all_sheets_text)
    except Exception as e:
        return f"Excelファイル {file_path} の読み込みに失敗しました: {str(e)}"

@tool
def update_excel(file_path: str, cell_values: dict[str, any]) -> str:
    """Update an Excel file with new data.
    **IMPORTANT**
    update_excelツールを同時並行で複数呼び出さないようにしてください。
    必ず一つのupdate_excelツールの実行が完了した後に更新結果を確認後、次のupdate_excelツールを呼び出してください。

    Args:
        file_path: Path to the Excel file
        cell_values: Dictionary mapping cell addresses to values.
                    セルアドレスは「シート名!セルアドレス」形式で指定可能（例: "Sheet1!A1"）
                    シート名を省略した場合はアクティブシートを対象とする（例: "A1"）
                    Example: {"Sheet1!A1": "Hello", "Sheet2!B2": 100, "C3": "World"}
    
    Returns:
        Success message with updated cells
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        updated_cells = []
        for cell_address, value in cell_values.items():
            # シート名!セルアドレス形式かどうかを判定
            if "!" in cell_address:
                sheet_name, cell_ref = cell_address.split("!", 1)
                if sheet_name not in wb.sheetnames:
                    return f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb.sheetnames}"
                ws = wb[sheet_name]
            else:
                ws = wb.active
                cell_ref = cell_address
            
            ws[cell_ref] = value
            updated_cells.append(f"{cell_address}={value}")
        
        wb.save(file_path)
        return f"Excel file updated successfully. Updated cells: {', '.join(updated_cells)}"
    except Exception as e:
        return f"Excelファイル {file_path} の更新に失敗しました: {str(e)} \n再度更新を実行してください。"

@tool
def read_result_csv(file_path: str, row_index: int) -> str:
    """Read a CSV file and return the value of the cell in the specified row
    
    Args:
        file_path: Path to the CSV file
        row_index: Index of the row to read(0-indexed)
    
    Returns:
        Value of the cell in the specified row
    """
    # 日本語CSVファイルの場合、cp932 または shift_jis を試す
    try:
        for encoding in ["utf-8", "cp932", "shift_jis"]:
            try:
                if row_index < 0 or row_index >= len(pd.read_csv(file_path, encoding=encoding)):
                    return f"指定された行が存在しません。max_row: {len(pd.read_csv(file_path, encoding=encoding))}"
                return pd.read_csv(file_path, encoding=encoding).iloc[row_index].to_string()
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Could not decode file {file_path} with any known encoding")
    except Exception as e:
        return f"Error reading file {file_path}: {str(e)}"

@tool
def validate_update(file_path: str) -> str:
    """Validate the update of the Excel file
    全シートを対象に「XXXX」と記載されているセルが更新されずに残っていないか確認する。
    条件に該当するセルは全て列挙して返す。
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        Validation result
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        not_updated_cells = []
        
        # 複数シートにXXXXがあるかどうかでフォーマット決定
        sheets_with_xxxx = set()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if any(cell == "XXXX" for cell in row if cell is not None):
                    sheets_with_xxxx.add(sheet_name)
                    break
        
        use_sheet_prefix = len(sheets_with_xxxx) > 1
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value == "XXXX":
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        if use_sheet_prefix:
                            cell_addr = f"{sheet_name}!{col_letter}{row_num}"
                        else:
                            cell_addr = f"{col_letter}{row_num}"
                        not_updated_cells.append(cell_addr)
        
        if not_updated_cells:
            return "以下のセルが更新されていません: " + ", ".join(not_updated_cells) + ". データを確認して更新してください。不明な場合は空白で更新してください。"
        return f"Excelファイル {file_path} の更新結果は正常です。（全{len(wb.sheetnames)}シート検証済み）"
    except Exception as e:
        return f"Excelファイル {file_path} の検証に失敗しました: {str(e)}"

file_path_result = "data/input/テストデータ_パターン2.csv"
file_path_result_length = len(pd.read_csv(file_path_result, encoding="shift_jis"))

# サブエージェントの定義
subagents = [
    {
        "name": "Sample_Result_Agent",
        "description": "各サンプルのテスト結果を抽出し整形するサブエージェントです。各サンプルのどのような情報を取得してほしいか具体的なカラム名などを指示してください。結果はJSON形式で返却されます。",
        "system_prompt": f"あなたは各サンプルに対して行われたテスト結果を管理する監査人です。指示された情報をJSON形式で返却してください。各サンプルテストの結果は{file_path_result}(len:{file_path_result_length})から取得してください。",
        "tools": [read_result_csv],
    },
    {
        "name": "Overview_Result_Agent",
        "description": "テスト結果全体を俯瞰して総括やメタデータを回答するサブエージェントです。希望する情報を渡すと総括やメタデータが回答されます。",
        "system_prompt": f"""
        あなたは監査人として職業的懐疑心を持ってテスト結果を総括しようとしています。
        指定された情報を{file_path_result}(len:{file_path_result_length})から取得して整理して回答してください。
        結果はJSON形式で返却してください。
        **判断基準**
        - サンプルごとのテストのうち、1件でも不備が見られたり、不備がないと言い切れない状況の場合は統制が有効であるとは言えません。
        - アクセスできるデータ（{file_path_result}）から回答できない情報は空白で返却してください。
        """,
        "tools": [read_result_csv],
    },
    ]


# %%
system_prompt = """
あなたは内部監査人として、Excelの調書を読み込んで更新しようとしています。サブエージェントを利用しながら以下のStepを完遂してください。

# Step1:
{file_path_template} のExcelのテンプレートをread_excelツールで読み込んで、記入すべき内容を特定してください。回答は不要です。
※ {file_path_example} のExcelの記入例もread_excelツールで読み込んで参考にしてください。

# Step2:
Step1で特定した記入すべき情報のうち、サンプルごとのテストに関連する情報を取得したい場合は
Sample_Result_Agentサブエージェントに依頼して情報を取得してください。
※サブエージェントは経緯やStep1の内容、Excelのテンプレートについては把握していないのでセルアドレスなどを渡さないでください。

# Step2.5:
Step2で取得した情報を整理して、update_excelツールを使って{file_path_template}に記入してください。
※記入内容の体裁や分量は{file_path_example}の記入例を参考にしてできるだけ踏襲してください。

# Step3:
Step1で特定した記入すべき情報のうち、Step2以外の情報を取得したい場合は
Overview_Result_Agentサブエージェントに依頼して情報を取得してください。
※サブエージェントは経緯やStep1の内容、Excelのテンプレートについては把握していないのでセルアドレスなどを渡さないでください。

# Step3.5:
Step2.5とStep3.5で取得した情報を整理して、update_excelツールを使って{file_path_template}に記入してください。
※記入内容の体裁や分量は{file_path_example}の記入例を参考にしてできるだけ踏襲してください。

# Step4:
validate_updateツールで更新結果を検証してください。結果が正常でない場合は必要に応じてStep2.5やStep3.5で情報を収集してください。

"""

# %%
agent = create_agent(
    model=llm_complex,
    tools=[read_excel, update_excel, validate_update],
    system_prompt=system_prompt,
    middleware=[
        SubAgentMiddleware(
            default_model=llm,
            default_tools=[read_excel, update_excel, read_result_csv, validate_update],
            subagents=subagents,
            general_purpose_agent=False,
        ),
    ],
    debug=True,
)

# %%
file_path_template = "data/input/複数シート型1_セクション別シート.xlsx"
file_path_example = "data/input/複数シート型1_セクション別シート_記入例.xlsx"
# %%
prompt = f"""
テンプレートのパス: {file_path_template}
記入例のパス: {file_path_example}
"""
# %%
inputs = {"messages": [{"role": "user", "content": prompt}]}
result = agent.invoke(inputs)
print(result.get("messages", [])[-1].content)
# %%
print_message_logs(extract_message_logs(result))


# %%
